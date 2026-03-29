"""
AGS Phillies — KPI Dashboard
A clean, professional management reporting dashboard.
Navy #001B4B | White | Black text.
Run: python app.py  (starts at http://127.0.0.1:8050)
"""

import os
import pandas as pd
import openpyxl
import dash
from dash import dcc, html, dash_table, Input, Output, State, callback
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime

# ─── CONFIG ──────────────────────────────────────────────────────────────────
# Resolve path relative to this file (works locally and on Render)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.environ.get("KPI_EXCEL_PATH", os.path.join(BASE_DIR, "kpi.xlsx"))
LOGO_URL = "https://res.cloudinary.com/dbviya1rj/image/upload/v1773384037/gpnkwelbdcwfjmw5axtx.webp"

# Design tokens
NAVY = "#001B4B"
DEEP_BLUE = "#1a5c9e"
WHITE = "#FFFFFF"
BLACK = "#000000"
LIGHT_GRAY = "#F4F4F4"
MID_GRAY = "#DCDCDC"
DARK_GRAY = "#555555"
BG = "#F0F0F0"


# ─── DATA LOADING ────────────────────────────────────────────────────────────
def load_sheets():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        hdrs = [str(h).strip() if h is not None else f"col_{i}"
                for i, h in enumerate(rows[0])]
        out[name] = pd.DataFrame([dict(zip(hdrs, r)) for r in rows[1:]
                                   if any(v is not None for v in r)])
    return out


def to_wlabel(val):
    if pd.isna(val): return "Unknown"
    return hasattr(val, "strftime") and val.strftime("W-%U") or "Unknown"


def h2m(t):
    """Hours/time to total minutes."""
    if t is None: return 0.0
    if isinstance(t, str):
        try:
            p = t.split(":"); return int(p[0])*60 + int(p[1]) + int(p[2])/60
        except: return 0.0
    if hasattr(t, "hour"): return t.hour*60 + t.minute + t.second/60
    return float(t)


raw = load_sheets()

# Transfer Rate
tdf = raw["agent_transfer_counts"].copy()
tdf["week_dt"] = pd.to_datetime(tdf["Week Ending"], errors="coerce")
tdf["week_label"] = tdf["week_dt"].apply(to_wlabel)
tdf = tdf.rename(columns={"agent":"agent","first_time_caller":"total_calls","transfer_count":"transfers"})
tdf["total_calls"] = pd.to_numeric(tdf["total_calls"], errors="coerce").fillna(0)
tdf["transfers"] = pd.to_numeric(tdf["transfers"], errors="coerce").fillna(0)
tdf["transfer_rate"] = (tdf["transfers"]/tdf["total_calls"].replace(0,1)*100).round(1)

# Admits
adf = raw["Admits"].copy()
adf["week_dt"] = pd.to_datetime(adf["Weekending"], errors="coerce")
adf["week_label"] = adf["week_dt"].apply(to_wlabel)
adf = adf.rename(columns={"User name":"agent","Admits":"admits"})
adf["admits"] = pd.to_numeric(adf["admits"], errors="coerce").fillna(0).astype(int)

# AHT
rdf = raw["AHT"].copy()
rdf["week_dt"] = pd.to_datetime(rdf["Weekending"], errors="coerce")
rdf["week_label"] = rdf["week_dt"].apply(to_wlabel)
rdf = rdf.rename(columns={"User name":"agent","Inbound calls":"inbound_calls",
                           "Inbound minutes":"inbound_minutes_raw","Hold time":"hold_time_raw"})
rdf["inbound_minutes"] = rdf["inbound_minutes_raw"].apply(h2m)
rdf["hold_minutes"] = rdf["hold_time_raw"].apply(h2m)
rdf["aht_minutes"] = (rdf["inbound_minutes"]/rdf["inbound_calls"].replace(0,1)).round(1)

# Attendance
atdf = raw["Attendance"].copy()
atdf["week_dt"] = pd.to_datetime(atdf["Week Ending"], errors="coerce")
atdf["week_label"] = atdf["week_dt"].apply(to_wlabel)
atdf = atdf.rename(columns={"agent":"agent","Total Hours Present":"hours_present",
                              "Total Hours Absent":"hours_absent","Attendance Percentage":"attendance_pct"})
atdf["attendance_pct"] = (pd.to_numeric(atdf["attendance_pct"],errors="coerce").fillna(0)*100).round(1)

# ─── MERGED ─────────────────────────────────────────────────────────────────
# Build merged with week_dt carried through for filtering
_merged_with_dt = (
    tdf[["week_label","week_dt","agent","total_calls","transfers","transfer_rate"]]
    .merge(adf[["week_label","agent","admits"]], on=["week_label","agent"], how="outer")
    .merge(rdf[["week_label","agent","aht_minutes","inbound_calls"]], on=["week_label","agent"], how="outer")
    .merge(atdf[["week_label","agent","attendance_pct"]], on=["week_label","agent"], how="outer")
)
# Fill numeric cols only; keep week_dt as-is (may be NaT)
numeric_cols = ["total_calls","transfers","transfer_rate","admits","aht_minutes","inbound_calls","attendance_pct"]
for c in numeric_cols:
    _merged_with_dt[c] = pd.to_numeric(_merged_with_dt[c], errors="coerce").fillna(0)
merged = _merged_with_dt.sort_values(["week_label","agent"]).reset_index(drop=True)

# Month options — computed from tdf which has clean week_dt before any merge
raw_months = (
    tdf.dropna(subset=["week_dt"])
    .assign(month=lambda d: d["week_dt"].dt.strftime("%Y-%m"))
    ["month"].unique().tolist()
)
MONTH_OPTIONS = [{"label": m, "value": m} for m in sorted(raw_months)]

# Agent-level KPI
akpi = (
    merged.groupby("agent").agg(
        avg_transfer_rate=("transfer_rate","mean"),
        total_admits=("admits","sum"),
        total_calls=("total_calls","sum"),
        total_transfers=("transfers","sum"),
        avg_aht=("aht_minutes","mean"),
        avg_attendance=("attendance_pct","mean"),
        weeks_active=("week_label","nunique"),
    ).reset_index()
)

# KPI Score: Attendance 25% | Transfer Rate 25% | Admits 35% | AHT 15%
def score(r):
    att = float(r.get("avg_attendance",0) or 0)
    trf = float(r.get("avg_transfer_rate",0) or 0)
    adm = float(r.get("total_admits",0) or 0)
    aht = float(r.get("avg_aht",0) or 0)
    wks = max(r.get("weeks_active",1) or 1, 1)
    adm_n = min(adm/wks/10*100, 100)
    aht_s = max(100-(aht/30*100), 0)
    return round(att*0.25 + trf*0.25 + adm_n*0.35 + aht_s*0.15, 1)

def pct_to_score(pct):
    """Map 0-100 percentage to 1-5 score scale."""
    if pct >= 90: return 5
    if pct >= 75: return 4
    if pct >= 60: return 3
    if pct >= 40: return 2
    return 1

akpi["kpi_score"] = akpi.apply(score, axis=1).apply(pct_to_score)
akpi["qa_score"] = 0  # QA data not yet available — pending source
akpi = akpi.sort_values("kpi_score", ascending=False).reset_index(drop=True)

WEEKS = sorted(merged["week_label"].unique())
WEEK_OPTIONS = [{"label": w, "value": w} for w in WEEKS]
N_AGENTS = len(akpi)
AVG_KPI = round(akpi["kpi_score"].mean(), 1)
TOP = akpi.iloc[0]
BOT = akpi.iloc[-1]

# ─── HELPERS ────────────────────────────────────────────────────────────────
def mc(label, val, sub="", highlight=False):
    """Metric card."""
    left = NAVY if highlight else MID_GRAY
    return html.Div([
        html.Div(label, style={
            "fontSize":"0.62rem","color":DARK_GRAY,
            "textTransform":"uppercase","letterSpacing":"1.4px",
            "fontWeight":"700","fontFamily":"Arial","marginBottom":"4px"
        }),
        html.Div(val, style={
            "fontSize":"1.6rem","fontWeight":"900","color":NAVY,
            "lineHeight":"1.1","fontFamily":"Arial Black,Arial"
        }),
        html.Div(sub, style={"fontSize":"0.68rem","color":DARK_GRAY,"marginTop":"4px","fontFamily":"Arial"}),
    ], style={
        "background":WHITE,"borderRadius":"8px","padding":"14px 18px",
        "boxShadow":"0 1px 4px rgba(0,0,0,0.07)",
        "border":f"1px solid {MID_GRAY}","borderLeft":f"4px solid {left}",
        "flex":"1","minWidth":"120px"
    })


def wrap(children, pad="20px"):
    return html.Div(children, style={
        "background":WHITE,"borderRadius":"10px","padding":pad,
        "boxShadow":"0 1px 4px rgba(0,0,0,0.07)",
        "border":f"1px solid {MID_GRAY}","margin":"0 32px"
    })


def hdr(text):
    return html.Div([
        html.Div(text, style={
            "fontSize":"0.72rem","fontWeight":"800","color":NAVY,
            "textTransform":"uppercase","letterSpacing":"1.2px",
            "fontFamily":"Arial Black,Arial","marginBottom":"2px"
        }),
        html.Div(style={"height":"3px","width":"28px","background":NAVY,"borderRadius":"2px"})
    ], style={"marginBottom":"12px"})


def gcf(fig):
    return dcc.Graph(figure=fig, config={"displayModeBar":False,
                                          "displaylogo":False,
                                          "modeBarButtonsToRemove":
                                          ["select2d","lasso2d","autoScale2d"]})


def tcolor(s):
    if s>=4: return DEEP_BLUE
    if s>=3: return NAVY
    if s>=2: return DARK_GRAY
    return "#AAAAAA"


# ─── CHARTS ─────────────────────────────────────────────────────────────────

# KPI BAR CHART
agents_s = [a.replace(" Phillies","") for a in akpi["agent"]]
fig_bar = go.Figure(go.Bar(
    y=agents_s[::-1], x=akpi["kpi_score"][::-1], orientation="h",
    marker_color=[tcolor(s) for s in akpi["kpi_score"][::-1]],
    text=[f"{s:.1f}" for s in akpi["kpi_score"][::-1]],
    textposition="outside",
    textfont=dict(size=10, color=BLACK, family="Arial"),
    hovertemplate="<b>%{y}</b><br>KPI: %{x:.1f}<extra></extra>",
    showlegend=False,
))
fig_bar.update_layout(
    title=dict(text="<b>Agent KPI Scores — Ranked</b>",
               font=dict(size=13, color=NAVY, family="Arial Black"), x=0.5, y=0.97),
    margin=dict(l=16, r=60, t=56, b=50),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=500,
    xaxis=dict(range=[0,5.5], tickvals=[1,2,3,4,5],
               ticktext=["1","2","3","4","5"],
               showgrid=True, gridcolor=MID_GRAY,
               title="KPI Score (1-5)", title_font=dict(size=10,color=DARK_GRAY),
               tickfont=dict(size=10,color=DARK_GRAY)),
    yaxis=dict(showgrid=False, tickfont=dict(size=10,color=BLACK), dtick=1),
    annotations=[
        dict(x=1,y=-1.0,text="1 Critical",showarrow=False,
             font=dict(size=8,color="#AAAAAA"),xanchor="center"),
        dict(x=2,y=-1.0,text="2 Needs Work",showarrow=False,
             font=dict(size=8,color=DARK_GRAY),xanchor="center"),
        dict(x=3,y=-1.0,text="3 Good",showarrow=False,
             font=dict(size=8,color=NAVY),xanchor="center"),
        dict(x=4.5,y=-1.0,text="4-5 Excellent",showarrow=False,
             font=dict(size=8,color=DEEP_BLUE),xanchor="center"),
    ],
    uniformtext_minsize=7, uniformtext_mode="show",
)


# WEEKLY TREND (3 metrics on one chart)
wk = merged.groupby("week_label").agg(
    att=("attendance_pct","mean"),
    trf=("transfer_rate","mean"),
    aht=("aht_minutes","mean"),
).reset_index().sort_values("week_label")

fig_trend = go.Figure()
fig_trend.add_trace(go.Scatter(
    x=wk["week_label"], y=wk["att"], mode="lines+markers+text",
    name="Attendance %", line=dict(color=DEEP_BLUE,width=2.5), marker=dict(size=8),
    text=[f"{v:.0f}%" for v in wk["att"]],
    textposition="top center", textfont=dict(size=9,color=DEEP_BLUE),
    yaxis="y1",
))
fig_trend.add_trace(go.Scatter(
    x=wk["week_label"], y=wk["trf"], mode="lines+markers+text",
    name="Transfer Rate %", line=dict(color=DARK_GRAY,width=2.5), marker=dict(size=8),
    text=[f"{v:.0f}%" for v in wk["trf"]],
    textposition="bottom center", textfont=dict(size=9,color=DARK_GRAY),
    yaxis="y2",
))
fig_trend.add_trace(go.Scatter(
    x=wk["week_label"], y=wk["aht"], mode="lines+markers+text",
    name="AHT (min)", line=dict(color="#999999",width=2.5), marker=dict(size=8),
    text=[f"{v:.1f}" for v in wk["aht"]],
    textposition="top center", textfont=dict(size=9,color="#999999"),
    yaxis="y3",
))
fig_trend.update_layout(
    title=dict(text="<b>Weekly KPI Trends</b>",
               font=dict(size=13,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
    margin=dict(l=50,r=70,t=56,b=50),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=300,
    xaxis=dict(showgrid=False,tickangle=-20,
               title="Week",title_font=dict(size=10,color=DARK_GRAY),
               tickfont=dict(size=10,color=BLACK)),
    yaxis=dict(title="Attendance %",title_font=dict(size=10,color=DEEP_BLUE),
               side="left",showgrid=True,gridcolor=MID_GRAY,
               tickfont=dict(size=10,color=DEEP_BLUE)),
    yaxis2=dict(title="Transfer %",title_font=dict(size=10,color=DARK_GRAY),
                overlaying="y",side="right",showgrid=False,
                tickfont=dict(size=10,color=DARK_GRAY)),
    yaxis3=dict(title="AHT min",title_font=dict(size=10,color="#999999"),
                overlaying="y",side="right",position=0.94,showgrid=False,
                tickfont=dict(size=10,color="#999999")),
    legend=dict(orientation="h",yanchor="bottom",y=1.09,
                xanchor="center",x=0.5,font=dict(size=10,color=BLACK),
                bgcolor="rgba(255,255,255,0.9)"),
    hovermode="x unified",
)


# ATTENDANCE HEATMAP (agents x weeks)
ap = merged.pivot_table(
    values="attendance_pct",index="agent",
    columns="week_label",aggfunc="mean"
).fillna(0)
agents_a = sorted(ap.index)
weeks_a = sorted(ap.columns)

fig_hm = go.Figure()
for i, w in enumerate(weeks_a):
    color = [NAVY, DEEP_BLUE, "#4A90D9"][i % 3]
    fig_hm.add_trace(go.Bar(
        name=w, x=[a.replace(" Phillies","") for a in agents_a], y=ap[w].reindex(agents_a).values,
        marker=dict(color=color, line_width=0),
        hovertemplate=f"Week {w}<br>%{{x}}<br>%:{{y:.1f}}%<extra></extra>",
        showlegend=True,
    ))

fig_hm.update_layout(
    title=dict(text="<b>Attendance % by Agent and Week</b>",
               font=dict(size=13,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
    margin=dict(l=50,r=30,t=56,b=120),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=420,
    barmode="group",
    xaxis=dict(tickangle=-25, showgrid=False,
               tickfont=dict(size=9,color=BLACK)),
    yaxis=dict(range=[0,110],showgrid=True,gridcolor=MID_GRAY,
               title="Attendance %",title_font=dict(size=10,color=DARK_GRAY),
               tickfont=dict(size=10,color=DARK_GRAY)),
    legend=dict(title="Week",orientation="h",yanchor="bottom",y=1.04,
                xanchor="center",x=0.5,font=dict(size=10,color=BLACK),
                bgcolor="rgba(255,255,255,0.9)"),
    uniformtext_minsize=7,
)


# KPI DISTRIBUTION PIE
dist_vals = [
    len(akpi[akpi["kpi_score"]>=5]),
    len(akpi[(akpi["kpi_score"]>=4)&(akpi["kpi_score"]<5)]),
    len(akpi[(akpi["kpi_score"]>=3)&(akpi["kpi_score"]<4)]),
    len(akpi[akpi["kpi_score"]<3]),
]
fig_pie = go.Figure(go.Pie(
    labels=["5 Excellent","4 Very Good","3 Good","1-2 Needs Work"],
    values=dist_vals,
    marker=dict(colors=[DEEP_BLUE,NAVY,DARK_GRAY,"#AAAAAA"]),
    textinfo="label+percent", textposition="outside",
    hole=0.45,
    hovertemplate="<b>%{label}</b><br>Count: %{value}<extra></extra>",
))
fig_pie.update_layout(
    margin=dict(l=20,r=20,t=20,b=20),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=240,
    showlegend=True,
    legend=dict(orientation="h",yanchor="bottom",y=-0.2,
                xanchor="center",x=0.5,font=dict(size=9,color=BLACK)),
    annotations=[dict(
        text=f"<b>{N_AGENTS}</b><br>Agents",x=0.5,y=0.5,showarrow=False,
        font=dict(size=12,color=NAVY)
    )],
)


# TRANSFER RATE BAR
trf_akpi = akpi.sort_values("avg_transfer_rate", ascending=False)
fig_trf = go.Figure(go.Bar(
    x=[a.replace(" Phillies","") for a in trf_akpi["agent"]],
    y=trf_akpi["avg_transfer_rate"],
    marker_color=[DARK_GRAY if v < 15 else "#AAAAAA" for v in trf_akpi["avg_transfer_rate"]],
    text=[f"{v:.1f}%" for v in trf_akpi["avg_transfer_rate"]],
    textposition="outside",
    textfont=dict(size=9,color=BLACK,family="Arial"),
    hovertemplate="<b>%{x}</b><br>Transfer Rate: %{y:.1f}%<extra></extra>",
    showlegend=False,
))
fig_trf.update_layout(
    title=dict(text="<b>Avg Transfer Rate % by Agent</b>  <span style='font-size:10px;color:#888'>(Target &lt;15% — gray bars are within target)</span>",
               font=dict(size=12,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
    margin=dict(l=40,r=20,t=56,b=80),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
    xaxis=dict(tickangle=-25,showgrid=False,tickfont=dict(size=9,color=BLACK)),
    yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
               title="Transfer Rate %",title_font=dict(size=10,color=DARK_GRAY),
               tickfont=dict(size=10,color=DARK_GRAY)),
    uniformtext_minsize=7,
)


# ADMITS BAR
adm_akpi = akpi.sort_values("total_admits", ascending=False)
fig_adm = go.Figure(go.Bar(
    x=[a.replace(" Phillies","") for a in adm_akpi["agent"]],
    y=adm_akpi["total_admits"],
    marker_color=DEEP_BLUE,
    text=[f"{int(v)}" for v in adm_akpi["total_admits"]],
    textposition="outside",
    textfont=dict(size=9,color=BLACK,family="Arial"),
    hovertemplate="<b>%{x}</b><br>Admits: %{y}<extra></extra>",
    showlegend=False,
))
fig_adm.update_layout(
    title=dict(text="<b>Total Admits by Agent</b>",
               font=dict(size=12,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
    margin=dict(l=40,r=20,t=56,b=80),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
    xaxis=dict(tickangle=-25,showgrid=False,tickfont=dict(size=9,color=BLACK)),
    yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
               title="Total Admits",title_font=dict(size=10,color=DARK_GRAY),
               tickfont=dict(size=10,color=DARK_GRAY)),
    uniformtext_minsize=7,
)


# AHT BAR
aht_akpi = akpi.sort_values("avg_aht")
fig_aht = go.Figure(go.Bar(
    x=[a.replace(" Phillies","") for a in aht_akpi["agent"]],
    y=aht_akpi["avg_aht"],
    marker_color=[NAVY if v <= 10 else DARK_GRAY if v <= 15 else "#AAAAAA"
                  for v in aht_akpi["avg_aht"]],
    text=[f"{v:.1f}" for v in aht_akpi["avg_aht"]],
    textposition="outside",
    textfont=dict(size=9,color=BLACK,family="Arial"),
    hovertemplate="<b>%{x}</b><br>AHT: %{y:.1f} min<extra></extra>",
    showlegend=False,
))
fig_aht.update_layout(
    title=dict(text="<b>Avg Handle Time (min) by Agent</b>  <span style='font-size:10px;color:#888'>(Navy=≤10, Gray=10-15, Light=&gt;15 min)</span>",
               font=dict(size=12,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
    margin=dict(l=40,r=20,t=56,b=80),
    plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
    xaxis=dict(tickangle=-25,showgrid=False,tickfont=dict(size=9,color=BLACK)),
    yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
               title="AHT (min)",title_font=dict(size=10,color=DARK_GRAY),
               tickfont=dict(size=10,color=DARK_GRAY)),
    uniformtext_minsize=7,
)


# ─── DATA TABLE ─────────────────────────────────────────────────────────────
tbl_df = akpi.copy()
tbl_df["agent"] = tbl_df["agent"].str.replace(" Phillies","")
tbl_df = tbl_df.round(1)

tbl = dash_table.DataTable(
    columns=[{"name":n,"id":i} for n,i in [
        ("Agent","agent"),("KPI Score","kpi_score"),
        ("QA Score","qa_score"),
        ("Attendance %","avg_attendance"),("Transfer Rate %","avg_transfer_rate"),
        ("Total Admits","total_admits"),("Avg AHT (min)","avg_aht"),
        ("Weeks Active","weeks_active"),
    ]],
    data=tbl_df.to_dict("records"),
    sort_action="native", filter_action="native", page_size=20,
    style_table={"overflowX":"auto"},
    style_header={
        "backgroundColor":NAVY,"color":WHITE,"fontWeight":"800",
        "fontSize":"0.72rem","textTransform":"uppercase",
        "letterSpacing":"0.4px","fontFamily":"Arial",
    },
    style_cell={
        "fontSize":"0.82rem","padding":"8px 12px",
        "fontFamily":"Arial","color":BLACK,
    },
    style_data_conditional=[
        {"if":{"row_index":"odd"},"backgroundColor":LIGHT_GRAY},
        {"if":{"column_id":"kpi_score","filter_query":"{kpi_score} >= 5"},
         "backgroundColor":"#D6E8F8","color":NAVY,"fontWeight":"700"},
        {"if":{"column_id":"kpi_score","filter_query":"{kpi_score} < 3"},
         "backgroundColor":"#E8E8E8","color":BLACK,"fontWeight":"700"},
    ],
)


# ─── HEADER ──────────────────────────────────────────────────────────────────
header = html.Div([
    html.Div([
        html.Img(src=LOGO_URL, style={
            "height":"42px","width":"auto","borderRadius":"6px","marginRight":"14px"
        }),
        html.Div([
            html.H1("AGS Phillies", style={
                "fontSize":"1.5rem","fontWeight":"900","color":WHITE,
                "margin":"0","fontFamily":"Arial Black,Arial"
            }),
            html.P("KPI Dashboard — Call Center Performance",
                   style={"color":"rgba(255,255,255,0.6)","margin":"1px 0 0",
                          "fontSize":"0.8rem","fontFamily":"Arial"})
        ]),
    ]),
    html.Div([
        html.Div([
            html.Span("LIVE", style={
                "color":NAVY,"background":WHITE,"fontSize":"0.62rem",
                "fontWeight":"800","padding":"2px 7px","borderRadius":"4px",
                "marginRight":"8px","letterSpacing":"0.5px"
            }),
            html.Span(datetime.now().strftime("%b %d, %Y"),
                     style={"color":"rgba(255,255,255,0.6)","fontSize":"0.78rem"}),
        ]),
        html.Span(f"{N_AGENTS} Agents  |  {len(WEEKS)} Weeks",
                 style={"color":"rgba(255,255,255,0.45)","fontSize":"0.73rem",
                        "marginTop":"3px","display":"block","textAlign":"right"}),
    ], style={"textAlign":"right"}),
], style={
    "background":NAVY,"padding":"18px 32px",
    "display":"flex","justifyContent":"space-between","alignItems":"center",
    "boxShadow":"0 4px 16px rgba(0,0,0,0.3)",
})


# ─── SUMMARY CARDS ───────────────────────────────────────────────────────────
top_ok = akpi[akpi["kpi_score"]>=5]
bot_bad = akpi[akpi["kpi_score"]<3]
sum_cards = html.Div([
    mc("Top Performer", TOP["agent"].replace(" Phillies",""),
       f"Score: {TOP['kpi_score']:.1f}", True),
    mc("Needs Attention", BOT["agent"].replace(" Phillies",""),
       f"Score: {BOT['kpi_score']:.1f}"),
    mc("Avg KPI Score", f"{AVG_KPI:.1f}", f"All {N_AGENTS} agents", True),
    mc("Excellent 5", len(top_ok), f"{len(top_ok)/N_AGENTS*100:.0f}% of team"),
    mc("Needs Work 1-2", len(bot_bad), f"{len(bot_bad)/N_AGENTS*100:.0f}% of team"),
    mc("Weeks Tracked", len(WEEKS), "Data coverage"),
], style={"display":"flex","gap":"12px","flexWrap":"wrap",
           "justifyContent":"center","padding":"16px 32px 0"})


# KPI bottom cards
overall_trf = merged["transfers"].sum()/merged["total_calls"].sum()*100
overall_adm = merged["admits"].sum()
overall_aht = merged["aht_minutes"].mean()
overall_att = merged["attendance_pct"].mean()

kpi_cards = html.Div([
    mc("Transfer Rate", f"{overall_trf:.1f}%", "Target <15%"),
    mc("Total Admits", f"{int(overall_adm)}", "All agents, all weeks"),
    mc("Avg Handle Time", f"{overall_aht:.1f} min", "Lower is better"),
    mc("Avg Attendance", f"{overall_att:.1f}%", "Target 95%+"),
], style={"display":"flex","gap":"12px","flexWrap":"wrap",
           "justifyContent":"center","padding":"0 32px"})


# ─── TAB LAYOUTS ─────────────────────────────────────────────────────────────
tab_kpi = html.Div([
    html.Div(style={"height":"16px"}),
    sum_cards,
    html.Div(style={"height":"16px"}),
    wrap([hdr("KPI Performance Score — All Agents Ranked"), gcf(fig_bar)]),
    html.Div(style={"height":"16px"}),
    kpi_cards,
    html.Div(style={"height":"16px"}),
    wrap([hdr("Full Agent KPI Breakdown"), tbl]),
    html.Div("⚠ QA Score data not yet available — pending source", style={
        "textAlign":"center","fontSize":"0.72rem","color":DARK_GRAY,
        "fontStyle":"italic","marginTop":"4px"
    }),
    html.Div(style={"height":"20px"}),
])

tab_charts = html.Div([
    html.Div(style={"height":"16px"}),
    sum_cards,
    html.Div(style={"height":"16px"}),

    # Two charts side by side: Trend + Pie
    html.Div([
        html.Div([
            wrap([hdr("Weekly KPI Trends"), gcf(fig_trend)])
        ], style={"flex":"1.5"}),
        html.Div([
            wrap([hdr("Score Distribution"), gcf(fig_pie)])
        ], style={"flex":"1"}),
    ], style={"display":"flex","gap":"0","padding":"0 0"}),

    html.Div(style={"height":"16px"}),

    # Attendance heatmap
    wrap([hdr("Attendance % by Agent and Week"), gcf(fig_hm)]),

    html.Div(style={"height":"16px"}),

    # Bottom 3 charts: Transfer, Admits, AHT
    html.Div([
        html.Div([wrap([gcf(fig_trf)])], style={"flex":"1"}),
        html.Div([wrap([gcf(fig_adm)])], style={"flex":"1"}),
        html.Div([wrap([gcf(fig_aht)])], style={"flex":"1"}),
    ], style={"display":"flex","gap":"0"}),

    html.Div(style={"height":"20px"}),
])


# ─── DYNAMIC FIGURES (used by callbacks) ───────────────────────────────────

def make_figures(f_m, f_akpi, tab):
    """Build all KPI / Charts tab figures from filtered merged & akpi dataframes."""
    # ── KPI bar ──────────────────────────────────────────────────────────────
    agents_s = [a.replace(" Phillies","") for a in f_akpi["agent"]]
    fig_bar = go.Figure(go.Bar(
        y=agents_s[::-1], x=f_akpi["kpi_score"][::-1], orientation="h",
        marker_color=[tcolor(s) for s in f_akpi["kpi_score"][::-1]],
        text=[f"{s:.1f}" for s in f_akpi["kpi_score"][::-1]],
        textposition="outside",
        textfont=dict(size=10, color=BLACK, family="Arial"),
        hovertemplate="<b>%{y}</b><br>KPI: %{x:.1f}<extra></extra>",
        showlegend=False,
    ))
    fig_bar.update_layout(
        title=dict(text="<b>Agent KPI Scores — Ranked</b>",
                   font=dict(size=13, color=NAVY, family="Arial Black"), x=0.5, y=0.97),
        margin=dict(l=16, r=60, t=56, b=50),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=500,
        xaxis=dict(range=[0,5.5], tickvals=[1,2,3,4,5],
                   ticktext=["1","2","3","4","5"],
                   showgrid=True, gridcolor=MID_GRAY,
                   title="KPI Score (1-5)", title_font=dict(size=10,color=DARK_GRAY),
                   tickfont=dict(size=10,color=DARK_GRAY)),
        yaxis=dict(showgrid=False, tickfont=dict(size=10,color=BLACK), dtick=1),
        annotations=[
            dict(x=1,y=-1.0,text="1 Critical",showarrow=False,
                 font=dict(size=8,color="#AAAAAA"),xanchor="center"),
            dict(x=2,y=-1.0,text="2 Needs Work",showarrow=False,
                 font=dict(size=8,color=DARK_GRAY),xanchor="center"),
            dict(x=3,y=-1.0,text="3 Good",showarrow=False,
                 font=dict(size=8,color=NAVY),xanchor="center"),
            dict(x=4.5,y=-1.0,text="4-5 Excellent",showarrow=False,
                 font=dict(size=8,color=DEEP_BLUE),xanchor="center"),
        ],
        uniformtext_minsize=7, uniformtext_mode="show",
    )

    # ── Summary cards ────────────────────────────────────────────────────────
    top_ok = f_akpi[f_akpi["kpi_score"]>=5]
    bot_bad = f_akpi[f_akpi["kpi_score"]<3]
    avg_kpi = round(f_akpi["kpi_score"].mean(), 1)
    top_row = f_akpi.iloc[0] if len(f_akpi) else None
    bot_row = f_akpi.iloc[-1] if len(f_akpi) else None
    sum_cards = html.Div([
        mc("Top Performer", top_row["agent"].replace(" Phillies","") if top_row is not None else "—",
           f"Score: {top_row['kpi_score']:.1f}" if top_row is not None else "—", True),
        mc("Needs Attention", bot_row["agent"].replace(" Phillies","") if bot_row is not None else "—",
           f"Score: {bot_row['kpi_score']:.1f}" if bot_row is not None else "—"),
        mc("Avg KPI Score", f"{avg_kpi:.1f}", f"All {len(f_akpi)} agents", True),
        mc("Excellent 5", len(top_ok), f"{len(top_ok)/max(len(f_akpi),1)*100:.0f}% of team"),
        mc("Needs Work 1-2", len(bot_bad), f"{len(bot_bad)/max(len(f_akpi),1)*100:.0f}% of team"),
        mc("Weeks Tracked", len(f_m["week_label"].unique()), "Data coverage"),
    ], style={"display":"flex","gap":"12px","flexWrap":"wrap",
               "justifyContent":"center","padding":"16px 32px 0"})

    # ── Table ────────────────────────────────────────────────────────────────
    tdf = f_akpi.copy()
    tdf["agent"] = tdf["agent"].str.replace(" Phillies","")
    tdf = tdf.round(1)
    tbl = dash_table.DataTable(
        columns=[{"name":n,"id":i} for n,i in [
            ("Agent","agent"),("KPI Score","kpi_score"),
            ("QA Score","qa_score"),
            ("Attendance %","avg_attendance"),("Transfer Rate %","avg_transfer_rate"),
            ("Total Admits","total_admits"),("Avg AHT (min)","avg_aht"),
            ("Weeks Active","weeks_active"),
        ]],
        data=tdf.to_dict("records"),
        sort_action="native", filter_action="native", page_size=20,
        style_table={"overflowX":"auto"},
        style_header={
            "backgroundColor":NAVY,"color":WHITE,"fontWeight":"800",
            "fontSize":"0.72rem","textTransform":"uppercase",
            "letterSpacing":"0.4px","fontFamily":"Arial",
        },
        style_cell={
            "fontSize":"0.82rem","padding":"8px 12px",
            "fontFamily":"Arial","color":BLACK,
        },
        style_data_conditional=[
            {"if":{"row_index":"odd"},"backgroundColor":LIGHT_GRAY},
            {"if":{"column_id":"kpi_score","filter_query":"{kpi_score} >= 5"},
             "backgroundColor":"#D6E8F8","color":NAVY,"fontWeight":"700"},
            {"if":{"column_id":"kpi_score","filter_query":"{kpi_score} < 3"},
             "backgroundColor":"#E8E8E8","color":BLACK,"fontWeight":"700"},
        ],
    )

    # ── KPI cards (overall metrics) ──────────────────────────────────────────
    overall_trf = f_m["transfers"].sum()/f_m["total_calls"].sum()*100 if f_m["total_calls"].sum() > 0 else 0
    overall_adm = f_m["admits"].sum()
    overall_aht = f_m["aht_minutes"].mean()
    overall_att = f_m["attendance_pct"].mean()
    kpi_cards = html.Div([
        mc("Transfer Rate", f"{overall_trf:.1f}%", "Target <15%"),
        mc("Total Admits", f"{int(overall_adm)}", "All agents, all weeks"),
        mc("Avg Handle Time", f"{overall_aht:.1f} min", "Lower is better"),
        mc("Avg Attendance", f"{overall_att:.1f}%", "Target 95%+"),
    ], style={"display":"flex","gap":"12px","flexWrap":"wrap",
               "justifyContent":"center","padding":"0 32px"})

    # ── Tab KPI content ──────────────────────────────────────────────────────
    if tab == "tab-kpi":
        return html.Div([
            html.Div(style={"height":"16px"}),
            sum_cards,
            html.Div(style={"height":"16px"}),
            wrap([hdr("KPI Performance Score — All Agents Ranked"), gcf(fig_bar)]),
            html.Div(style={"height":"16px"}),
            kpi_cards,
            html.Div(style={"height":"16px"}),
            wrap([hdr("Full Agent KPI Breakdown"), tbl]),
            html.Div("⚠ QA Score data not yet available — pending source", style={
                "textAlign":"center","fontSize":"0.72rem","color":DARK_GRAY,
                "fontStyle":"italic","marginTop":"4px"
            }),
            html.Div(style={"height":"20px"}),
        ])

    # ── Charts tab ────────────────────────────────────────────────────────────
    n_agents = len(f_akpi)

    # Weekly trend
    wk = f_m.groupby("week_label").agg(
        att=("attendance_pct","mean"),
        trf=("transfer_rate","mean"),
        aht=("aht_minutes","mean"),
    ).reset_index().sort_values("week_label")
    fig_trend = go.Figure()
    fig_trend.add_trace(go.Scatter(
        x=wk["week_label"], y=wk["att"], mode="lines+markers+text",
        name="Attendance %", line=dict(color=DEEP_BLUE,width=2.5), marker=dict(size=8),
        text=[f"{v:.0f}%" for v in wk["att"]],
        textposition="top center", textfont=dict(size=9,color=DEEP_BLUE),
        yaxis="y1",
    ))
    fig_trend.add_trace(go.Scatter(
        x=wk["week_label"], y=wk["trf"], mode="lines+markers+text",
        name="Transfer Rate %", line=dict(color=DARK_GRAY,width=2.5), marker=dict(size=8),
        text=[f"{v:.0f}%" for v in wk["trf"]],
        textposition="bottom center", textfont=dict(size=9,color=DARK_GRAY),
        yaxis="y2",
    ))
    fig_trend.add_trace(go.Scatter(
        x=wk["week_label"], y=wk["aht"], mode="lines+markers+text",
        name="AHT (min)", line=dict(color="#999999",width=2.5), marker=dict(size=8),
        text=[f"{v:.1f}" for v in wk["aht"]],
        textposition="top center", textfont=dict(size=9,color="#999999"),
        yaxis="y3",
    ))
    fig_trend.update_layout(
        title=dict(text="<b>Weekly KPI Trends</b>",
                   font=dict(size=13,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
        margin=dict(l=50,r=70,t=56,b=50),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=300,
        xaxis=dict(showgrid=False,tickangle=-20,
                   title="Week",title_font=dict(size=10,color=DARK_GRAY),
                   tickfont=dict(size=10,color=BLACK)),
        yaxis=dict(title="Attendance %",title_font=dict(size=10,color=DEEP_BLUE),
                   side="left",showgrid=True,gridcolor=MID_GRAY,
                   tickfont=dict(size=10,color=DEEP_BLUE)),
        yaxis2=dict(title="Transfer %",title_font=dict(size=10,color=DARK_GRAY),
                    overlaying="y",side="right",showgrid=False,
                    tickfont=dict(size=10,color=DARK_GRAY)),
        yaxis3=dict(title="AHT min",title_font=dict(size=10,color="#999999"),
                    overlaying="y",side="right",position=0.94,showgrid=False,
                    tickfont=dict(size=10,color="#999999")),
        legend=dict(orientation="h",yanchor="bottom",y=1.09,
                    xanchor="center",x=0.5,font=dict(size=10,color=BLACK),
                    bgcolor="rgba(255,255,255,0.9)"),
        hovermode="x unified",
    )

    # Distribution pie
    dist_vals = [
        len(f_akpi[f_akpi["kpi_score"]>=5]),
        len(f_akpi[(f_akpi["kpi_score"]>=4)&(f_akpi["kpi_score"]<5)]),
        len(f_akpi[(f_akpi["kpi_score"]>=3)&(f_akpi["kpi_score"]<4)]),
        len(f_akpi[f_akpi["kpi_score"]<3]),
    ]
    fig_pie = go.Figure(go.Pie(
        labels=["5 Excellent","4 Very Good","3 Good","1-2 Needs Work"],
        values=dist_vals,
        marker=dict(colors=[DEEP_BLUE,NAVY,DARK_GRAY,"#AAAAAA"]),
        textinfo="label+percent", textposition="outside",
        hole=0.45,
        hovertemplate="<b>%{label}</b><br>Count: %{value}<extra></extra>",
    ))
    fig_pie.update_layout(
        margin=dict(l=20,r=20,t=20,b=20),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=240,
        showlegend=True,
        legend=dict(orientation="h",yanchor="bottom",y=-0.2,
                    xanchor="center",x=0.5,font=dict(size=9,color=BLACK)),
        annotations=[dict(
            text=f"<b>{n_agents}</b><br>Agents",x=0.5,y=0.5,showarrow=False,
            font=dict(size=12,color=NAVY)
        )],
    )

    # Attendance heatmap
    ap = f_m.pivot_table(
        values="attendance_pct",index="agent",
        columns="week_label",aggfunc="mean"
    ).fillna(0)
    agents_a = sorted(ap.index)
    weeks_a = sorted(ap.columns)
    fig_hm = go.Figure()
    for i, w in enumerate(weeks_a):
        color = [NAVY, DEEP_BLUE, "#4A90D9"][i % 3]
        fig_hm.add_trace(go.Bar(
            name=w, x=[a.replace(" Phillies","") for a in agents_a], y=ap[w].reindex(agents_a).values,
            marker=dict(color=color, line_width=0),
            hovertemplate=f"Week {w}<br>%{{x}}<br>%:{{y:.1f}}%<extra></extra>",
            showlegend=True,
        ))
    fig_hm.update_layout(
        title=dict(text="<b>Attendance % by Agent and Week</b>",
                   font=dict(size=13,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
        margin=dict(l=50,r=30,t=56,b=120),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=420,
        barmode="group",
        xaxis=dict(tickangle=-25, showgrid=False,
                   tickfont=dict(size=9,color=BLACK)),
        yaxis=dict(range=[0,110],showgrid=True,gridcolor=MID_GRAY,
                   title="Attendance %",title_font=dict(size=10,color=DARK_GRAY),
                   tickfont=dict(size=10,color=DARK_GRAY)),
        legend=dict(title="Week",orientation="h",yanchor="bottom",y=1.04,
                    xanchor="center",x=0.5,font=dict(size=10,color=BLACK),
                    bgcolor="rgba(255,255,255,0.9)"),
        uniformtext_minsize=7,
    )

    # Transfer rate bar
    trf_akpi = f_akpi.sort_values("avg_transfer_rate", ascending=False)
    fig_trf = go.Figure(go.Bar(
        x=[a.replace(" Phillies","") for a in trf_akpi["agent"]],
        y=trf_akpi["avg_transfer_rate"],
        marker_color=[DARK_GRAY if v < 15 else "#AAAAAA" for v in trf_akpi["avg_transfer_rate"]],
        text=[f"{v:.1f}%" for v in trf_akpi["avg_transfer_rate"]],
        textposition="outside",
        textfont=dict(size=9,color=BLACK,family="Arial"),
        hovertemplate="<b>%{x}</b><br>Transfer Rate: %{y:.1f}%<extra></extra>",
        showlegend=False,
    ))
    fig_trf.update_layout(
        title=dict(text="<b>Avg Transfer Rate % by Agent</b>  <span style='font-size:10px;color:#888'>(Target &lt;15% — gray bars are within target)</span>",
                   font=dict(size=12,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
        margin=dict(l=40,r=20,t=56,b=80),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
        xaxis=dict(tickangle=-25,showgrid=False,tickfont=dict(size=9,color=BLACK)),
        yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
                   title="Transfer Rate %",title_font=dict(size=10,color=DARK_GRAY),
                   tickfont=dict(size=10,color=DARK_GRAY)),
        uniformtext_minsize=7,
    )

    # Admits bar
    adm_akpi = f_akpi.sort_values("total_admits", ascending=False)
    fig_adm = go.Figure(go.Bar(
        x=[a.replace(" Phillies","") for a in adm_akpi["agent"]],
        y=adm_akpi["total_admits"],
        marker_color=DEEP_BLUE,
        text=[f"{int(v)}" for v in adm_akpi["total_admits"]],
        textposition="outside",
        textfont=dict(size=9,color=BLACK,family="Arial"),
        hovertemplate="<b>%{x}</b><br>Admits: %{y}<extra></extra>",
        showlegend=False,
    ))
    fig_adm.update_layout(
        title=dict(text="<b>Total Admits by Agent</b>",
                   font=dict(size=12,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
        margin=dict(l=40,r=20,t=56,b=80),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
        xaxis=dict(tickangle=-25,showgrid=False,tickfont=dict(size=9,color=BLACK)),
        yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
                   title="Total Admits",title_font=dict(size=10,color=DARK_GRAY),
                   tickfont=dict(size=10,color=DARK_GRAY)),
        uniformtext_minsize=7,
    )

    # AHT bar
    aht_akpi = f_akpi.sort_values("avg_aht")
    fig_aht = go.Figure(go.Bar(
        x=[a.replace(" Phillies","") for a in aht_akpi["agent"]],
        y=aht_akpi["avg_aht"],
        marker_color=[NAVY if v <= 10 else DARK_GRAY if v <= 15 else "#AAAAAA"
                      for v in aht_akpi["avg_aht"]],
        text=[f"{v:.1f}" for v in aht_akpi["avg_aht"]],
        textposition="outside",
        textfont=dict(size=9,color=BLACK,family="Arial"),
        hovertemplate="<b>%{x}</b><br>AHT: %{y:.1f} min<extra></extra>",
        showlegend=False,
    ))
    fig_aht.update_layout(
        title=dict(text="<b>Avg Handle Time (min) by Agent</b>  <span style='font-size:10px;color:#888'>(Navy=≤10, Gray=10-15, Light=&gt;15 min)</span>",
                   font=dict(size=12,color=NAVY,family="Arial Black"),x=0.5,y=0.97),
        margin=dict(l=40,r=20,t=56,b=80),
        plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
        xaxis=dict(tickangle=-25,showgrid=False,tickfont=dict(size=9,color=BLACK)),
        yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
                   title="AHT (min)",title_font=dict(size=10,color=DARK_GRAY),
                   tickfont=dict(size=10,color=DARK_GRAY)),
        uniformtext_minsize=7,
    )

    # ── Charts tab content ───────────────────────────────────────────────────
    return html.Div([
        html.Div(style={"height":"16px"}),
        sum_cards,
        html.Div(style={"height":"16px"}),
        html.Div([
            html.Div([wrap([hdr("Weekly KPI Trends"), gcf(fig_trend)])], style={"flex":"1.5"}),
            html.Div([wrap([hdr("Score Distribution"), gcf(fig_pie)])], style={"flex":"1"}),
        ], style={"display":"flex","gap":"0","padding":"0 0"}),
        html.Div(style={"height":"16px"}),
        wrap([hdr("Attendance % by Agent and Week"), gcf(fig_hm)]),
        html.Div(style={"height":"16px"}),
        html.Div([
            html.Div([wrap([gcf(fig_trf)])], style={"flex":"1"}),
            html.Div([wrap([gcf(fig_adm)])], style={"flex":"1"}),
            html.Div([wrap([gcf(fig_aht)])], style={"flex":"1"}),
        ], style={"display":"flex","gap":"0"}),
        html.Div(style={"height":"20px"}),
    ])


# ─── WEEK-OVER-WEEK COMPARISON ──────────────────────────────────────────────

def make_compare_figures(week_a, week_b, f_merged, f_akpi):
    """Build 4 week-over-week comparison charts."""
    if not week_a or not week_b:
        return [html.Div("Select two weeks to compare", style={"textAlign":"center","color":DARK_GRAY,"padding":"40px"})] * 4

    def week_df(w):
        return f_merged[f_merged["week_label"] == w]

    def week_agent_df(w):
        return f_akpi.copy()

    wdf_a = week_df(week_a)
    wdf_b = week_df(week_b)
    agents_all = sorted(set(wdf_a["agent"].unique()) | set(wdf_b["agent"].unique()))

    def compare_trace(metric, label, color_a, color_b):
        agents_a_s = [a.replace(" Phillies","") for a in agents_all]
        vals_a = [wdf_a[wdf_a["agent"]==a][metric].mean() if len(wdf_a[wdf_a["agent"]==a]) else 0 for a in agents_all]
        vals_b = [wdf_b[wdf_b["agent"]==a][metric].mean() if len(wdf_b[wdf_b["agent"]==a]) else 0 for a in agents_all]
        fig = go.Figure()
        fig.add_trace(go.Bar(name=week_a, x=agents_a_s, y=vals_a, marker_color=color_a,
                             text=[f"{v:.1f}" if v else "" for v in vals_a], textposition="outside",
                             hovertemplate=f"<b>{{x}}</b><br>{label}: %{{y:.1f}}<extra></extra>"))
        fig.add_trace(go.Bar(name=week_b, x=agents_a_s, y=vals_b, marker_color=color_b,
                             text=[f"{v:.1f}" if v else "" for v in vals_b], textposition="outside",
                             hovertemplate=f"<b>{{x}}</b><br>{label}: %{{y:.1f}}<extra></extra>"))
        fig.update_layout(
            barmode="group", plot_bgcolor=WHITE, paper_bgcolor=WHITE, height=280,
            margin=dict(l=40,r=20,t=50,b=80),
            xaxis=dict(tickangle=-20, showgrid=False, tickfont=dict(size=9,color=BLACK)),
            yaxis=dict(showgrid=True,gridcolor=MID_GRAY,
                       title=label,title_font=dict(size=10,color=DARK_GRAY),
                       tickfont=dict(size=10,color=DARK_GRAY)),
            legend=dict(font=dict(size=10,color=BLACK)),
            uniformtext_minsize=7,
        )
        return fig

    return [
        gcf(compare_trace("attendance_pct", "Attendance %", DEEP_BLUE, "#4A90D9")),
        gcf(compare_trace("transfer_rate", "Transfer Rate %", DARK_GRAY, "#AAAAAA")),
        gcf(compare_trace("aht_minutes", "AHT (min)", "#999999", "#CCCCCC")),
        gcf(compare_trace("admits", "Admits", NAVY, DEEP_BLUE)),
    ]


# ─── FOOTER ─────────────────────────────────────────────────────────────────
footer = html.Div([
    html.Span([
        "AGS Phillies KPI Dashboard",
        f"  |  Updated: {datetime.now().strftime('%b %d, %Y %H:%M')}",
        f"  |  Avg KPI: {AVG_KPI:.1f}/5",
        f"  |  {N_AGENTS} Agents  |  {len(WEEKS)} Weeks",
    ], style={"color":"rgba(255,255,255,0.5)","fontSize":"0.72rem",
              "textAlign":"center","display":"block"})
], style={"background":NAVY,"padding":"12px 32px","marginTop":"8px"})


# ─── APP ─────────────────────────────────────────────────────────────────────
app = dash.Dash(__name__, title="AGS Phillies KPI Dashboard",
                suppress_callback_exceptions=True)

# Filter bar (above tabs)
filter_bar = html.Div([
    html.Div([
        html.Span("Filters:", style={
            "fontWeight":"700","color":NAVY,"marginRight":"8px",
            "fontSize":"0.78rem","fontFamily":"Arial","alignSelf":"center"
        }),
        dcc.Dropdown(
            id="month-filter",
            options=[{"label":"All Months","value":""}] + MONTH_OPTIONS,
            value="", placeholder="All Months",
            style={"width":"180px","fontSize":"0.78rem","marginRight":"12px"},
            clearable=False,
        ),
        dcc.Dropdown(
            id="week-filter",
            options=[{"label":"All Weeks","value":""}] + WEEK_OPTIONS,
            value="", placeholder="All Weeks",
            style={"width":"150px","fontSize":"0.78rem","marginRight":"12px"},
            clearable=False,
        ),
        # Week A / B for comparison
        html.Span("Compare:", style={
            "fontWeight":"700","color":NAVY,"marginRight":"8px",
            "fontSize":"0.78rem","fontFamily":"Arial","alignSelf":"center",
            "marginLeft":"12px"
        }),
        dcc.Dropdown(
            id="week-a",
            options=WEEK_OPTIONS,
            value=WEEKS[-2] if len(WEEKS) >= 2 else (WEEKS[-1] if WEEKS else ""),
            placeholder="Week A", style={"width":"130px","fontSize":"0.78rem","marginRight":"8px"},
        ),
        dcc.Dropdown(
            id="week-b",
            options=WEEK_OPTIONS,
            value=WEEKS[-1] if WEEKS else "",
            placeholder="Week B", style={"width":"130px","fontSize":"0.78rem"},
        ),
    ], style={
        "display":"flex","alignItems":"center","flexWrap":"wrap",
        "padding":"10px 32px","background":WHITE,
        "borderBottom":f"1px solid {MID_GRAY}"
    }),
], style={"background":BG})

# Hidden stores
store_filter = dcc.Store(id="store-filter", data={
    "merged": merged.to_dict("records"),
    "akpi": akpi.to_dict("records"),
})
store_compare = dcc.Store(id="store-compare", data={"week_a":"","week_b":""})

app.layout = html.Div([
    header,
    filter_bar,
    html.Div([
        dcc.Tabs(
            id="tabs", value="tab-kpi",
            children=[
                dcc.Tab(label="KPI SCORE", value="tab-kpi",
                    style={"fontFamily":"Arial","fontWeight":"700","fontSize":"0.85rem",
                           "padding":"12px 28px","borderBottom":"3px solid transparent",
                           "color":DARK_GRAY,"background":BG},
                    selected_style={"fontFamily":"Arial","fontWeight":"800","fontSize":"0.85rem",
                                   "padding":"12px 28px","borderBottom":f"3px solid {NAVY}",
                                   "color":NAVY,"background":WHITE},
                ),
                dcc.Tab(label="CHARTS & TRENDS", value="tab-charts",
                    style={"fontFamily":"Arial","fontWeight":"700","fontSize":"0.85rem",
                           "padding":"12px 28px","borderBottom":"3px solid transparent",
                           "color":DARK_GRAY,"background":BG},
                    selected_style={"fontFamily":"Arial","fontWeight":"800","fontSize":"0.85rem",
                                   "padding":"12px 28px","borderBottom":f"3px solid {NAVY}",
                                   "color":NAVY,"background":WHITE},
                ),
                dcc.Tab(label="COMPARE WEEKS", value="tab-compare",
                    style={"fontFamily":"Arial","fontWeight":"700","fontSize":"0.85rem",
                           "padding":"12px 28px","borderBottom":"3px solid transparent",
                           "color":DARK_GRAY,"background":BG},
                    selected_style={"fontFamily":"Arial","fontWeight":"800","fontSize":"0.85rem",
                                   "padding":"12px 28px","borderBottom":f"3px solid {NAVY}",
                                   "color":NAVY,"background":WHITE},
                ),
            ],
            style={"margin":"0 32px","borderBottom":f"2px solid {MID_GRAY}","background":BG},
        ),
    ]),
    html.Div(id="tab-content", style={"background":BG,"minHeight":"85vh","paddingBottom":"20px"}),
    store_filter,
    store_compare,
    footer,
], style={"background":BG,"fontFamily":"Arial"})


# ─── CALLBACKS ───────────────────────────────────────────────────────────────

def _compute_akpi(df):
    """Recompute agent KPI from a filtered merged dataframe."""
    a = df.groupby("agent").agg(
        avg_transfer_rate=("transfer_rate","mean"),
        total_admits=("admits","sum"),
        total_calls=("total_calls","sum"),
        total_transfers=("transfers","sum"),
        avg_aht=("aht_minutes","mean"),
        avg_attendance=("attendance_pct","mean"),
        weeks_active=("week_label","nunique"),
    ).reset_index()
    a["qa_score"] = 0
    a["kpi_score"] = a.apply(score, axis=1).apply(pct_to_score)
    return a.sort_values("kpi_score", ascending=False).reset_index(drop=True)


@app.callback(
    Output("store-filter", "data"),
    Input("month-filter", "value"),
    Input("week-filter", "value"),
)
def update_store_filter(month_val, week_val):
    df = merged.copy()
    if month_val:
        df = df[pd.to_datetime(df["week_dt"]).dt.strftime("%Y-%m") == month_val]
    if week_val:
        df = df[df["week_label"] == week_val]
    return {"merged": df.to_dict("records"), "akpi": _compute_akpi(df).to_dict("records")}


@app.callback(
    Output("tab-content", "children"),
    Input("store-filter", "data"),
    Input("store-compare", "data"),
    State("tabs", "value"),
    State("month-filter", "value"),
    State("week-filter", "value"),
)
def update_tab_content(store_data, compare_data, tab, month_val, week_val):
    f_merged = pd.DataFrame.from_dict(store_data["merged"])
    f_akpi = pd.DataFrame.from_dict(store_data["akpi"])

    if tab == "tab-compare":
        week_a = compare_data.get("week_a") or ""
        week_b = compare_data.get("week_b") or ""
        compare_figs = make_compare_figures(week_a, week_b, f_merged, f_akpi)
        return html.Div([
            html.Div(style={"height":"16px"}),
            wrap([hdr(f"Week-over-Week Comparison: {week_a or '?'} vs {week_b or '?'}")]),
            html.Div(style={"height":"8px"}),
            html.Div([
                html.Div([wrap([hdr("Attendance %"), compare_figs[0]])], style={"flex":"1"}),
                html.Div([wrap([hdr("Transfer Rate %"), compare_figs[1]])], style={"flex":"1"}),
            ], style={"display":"flex","gap":"0"}),
            html.Div(style={"height":"12px"}),
            html.Div([
                html.Div([wrap([hdr("Avg Handle Time"), compare_figs[2]])], style={"flex":"1"}),
                html.Div([wrap([hdr("Total Admits"), compare_figs[3]])], style={"flex":"1"}),
            ], style={"display":"flex","gap":"0"}),
            html.Div(style={"height":"20px"}),
        ])

    return make_figures(f_merged, f_akpi, tab)


@app.callback(
    Output("store-compare", "data"),
    Input("week-a", "value"),
    Input("week-b", "value"),
)
def update_store_compare(week_a, week_b):
    return {"week_a": week_a or "", "week_b": week_b or ""}


if __name__=="__main__":
    print("="*55)
    print("  AGS Phillies KPI Dashboard")
    print(f"  Agents : {N_AGENTS}")
    print(f"  Weeks  : {len(WEEKS)}")
    print(f"  Avg KPI: {AVG_KPI:.1f}/5")
    print(f"  Top    : {TOP['agent']} ({TOP['kpi_score']:.1f})")
    print(f"  Bottom : {BOT['agent']} ({BOT['kpi_score']:.1f})")
    print("  Open   : http://127.0.0.1:8050")
    print("="*55)
    app.run(debug=True, port=8050, host="127.0.0.1")
