"""
AGS Phillies KPI Dashboard — Static Image Generator
Generates a high-quality PNG dashboard image for management reporting.
Run: python generate_image.py
Output: ags_kpi_dashboard.png in the same directory
"""

import pandas as pd
import openpyxl
from datetime import datetime
import os

# ─── Load & Process Data ─────────────────────────────────────────────────────
EXCEL_PATH = "/Users/archerterminez/Desktop/AGS/Dashboard/kpi.xlsx"

def load_sheets():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(rows[0])]
        data = [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]
        sheets[name] = pd.DataFrame(data)
    return sheets

def hours_to_minutes(t):
    if t is None: return 0.0
    if isinstance(t, str):
        try:
            parts = t.split(":")
            return int(parts[0]) * 60 + int(parts[1]) + int(parts[2]) / 60
        except: return 0.0
    if hasattr(t, "hour"):
        return t.hour * 60 + t.minute + t.second / 60
    return float(t)

raw = load_sheets()

# Transfer Rate
transfer_df = raw["agent_transfer_counts"].copy()
if "Week Ending" in transfer_df.columns:
    transfer_df["Week Ending"] = pd.to_datetime(transfer_df["Week Ending"], errors="coerce")
transfer_df = transfer_df.rename(columns={
    "agent": "agent", "first_time_caller": "total_calls",
    "transfer_count": "transfers"
})
transfer_df["total_calls"] = pd.to_numeric(transfer_df["total_calls"], errors="coerce").fillna(0)
transfer_df["transfers"] = pd.to_numeric(transfer_df["transfers"], errors="coerce").fillna(0)
transfer_df["transfer_rate"] = (transfer_df["transfers"] / transfer_df["total_calls"].replace(0, 1) * 100).round(1)
transfer_df["week_label"] = transfer_df["Week Ending"].apply(
    lambda x: x.strftime("W-%U") if pd.notna(x) else "Unknown")

# Admits
admits_df = raw["Admits"].copy()
if "Weekending" in admits_df.columns:
    admits_df["Weekending"] = pd.to_datetime(admits_df["Weekending"], errors="coerce")
admits_df = admits_df.rename(columns={"User name": "agent", "Admits": "admits"})
admits_df["admits"] = pd.to_numeric(admits_df["admits"], errors="coerce").fillna(0).astype(int)
admits_df["week_label"] = admits_df["Weekending"].apply(
    lambda x: x.strftime("W-%U") if pd.notna(x) else "Unknown")

# AHT
aht_df = raw["AHT"].copy()
if "Weekending" in aht_df.columns:
    aht_df["Weekending"] = pd.to_datetime(aht_df["Weekending"], errors="coerce")
aht_df = aht_df.rename(columns={
    "User name": "agent", "Inbound calls": "inbound_calls",
    "Inbound minutes": "inbound_minutes_raw", "Hold time": "hold_time_raw"
})
aht_df["inbound_minutes"] = aht_df["inbound_minutes_raw"].apply(hours_to_minutes)
aht_df["hold_minutes"] = aht_df["hold_time_raw"].apply(hours_to_minutes)
aht_df["aht_minutes"] = (aht_df["inbound_minutes"] / aht_df["inbound_calls"].replace(0, 1)).round(1)
aht_df["week_label"] = aht_df["Weekending"].apply(
    lambda x: x.strftime("W-%U") if pd.notna(x) else "Unknown")

# Attendance
attend_df = raw["Attendance"].copy()
if "Week Ending" in attend_df.columns:
    attend_df["Week Ending"] = pd.to_datetime(attend_df["Week Ending"], errors="coerce")
attend_df = attend_df.rename(columns={
    "agent": "agent", "Total Hours Present": "hours_present",
    "Total Hours Absent": "hours_absent", "Attendance Percentage": "attendance_pct"
})
attend_df["attendance_pct"] = pd.to_numeric(attend_df["attendance_pct"], errors="coerce").fillna(0) * 100
attend_df["week_label"] = attend_df["Week Ending"].apply(
    lambda x: x.strftime("W-%U") if pd.notna(x) else "Unknown")

# Merge
merged = (
    transfer_df[["week_label", "agent", "total_calls", "transfers", "transfer_rate"]]
    .merge(admits_df[["week_label", "agent", "admits"]], on=["week_label", "agent"], how="outer")
    .merge(aht_df[["week_label", "agent", "aht_minutes", "inbound_calls"]], on=["week_label", "agent"], how="outer")
    .merge(attend_df[["week_label", "agent", "attendance_pct"]], on=["week_label", "agent"], how="outer")
    .fillna(0)
)
merged = merged.sort_values(["week_label", "agent"]).reset_index(drop=True)

# Agent KPI
agent_kpi = (
    merged.groupby("agent")
    .agg(
        avg_transfer_rate=("transfer_rate", "mean"),
        total_admits=("admits", "sum"),
        avg_aht=("aht_minutes", "mean"),
        avg_attendance=("attendance_pct", "mean"),
        weeks_active=("week_label", "nunique"),
    )
    .reset_index()
)

def calc_kpi(row):
    attendance = float(row.get("avg_attendance", 0) or 0)
    transfer = float(row.get("avg_transfer_rate", 0) or 0)
    admits = float(row.get("total_admits", 0) or 0)
    aht = float(row.get("avg_aht", 0) or 0)
    admits_norm = min(admits / (row.get("weeks_active", 1) or 1) / 10 * 100, 100)
    score = attendance * 0.25 + transfer * 0.25 + admits_norm * 0.35 + (100 - min(aht, 30) / 30 * 100) * 0.15
    return round(score, 1)

agent_kpi["kpi_score"] = agent_kpi.apply(calc_kpi, axis=1)
agent_kpi = agent_kpi.sort_values("kpi_score", ascending=False).reset_index(drop=True)

# ─── Colors ──────────────────────────────────────────────────────────────────
C = {
    "primary": "#0F2847",
    "secondary": "#1A4A7A",
    "accent": "#00C9A7",
    "accent2": "#7B61FF",
    "good": "#22C55E",
    "warn": "#F59E0B",
    "bad": "#EF4444",
    "tier_excellent": "#0066CC",
    "tier_good": "#22C55E",
    "tier_mid": "#F59E0B",
    "tier_low": "#EF4444",
    "bg": "#F0F4F8",
    "text": "#1E293B",
    "muted": "#64748B",
}

def tier_color(s):
    if s >= 80: return C["tier_excellent"]
    if s >= 60: return C["tier_good"]
    if s >= 40: return C["tier_mid"]
    return C["tier_low"]

# ─── Build Chart Figure ───────────────────────────────────────────────────────
import plotly.graph_objects as go
from plotly.subplots import make_subplots

avg_kpi = agent_kpi["kpi_score"].mean()
top_agent = agent_kpi.iloc[0]
bottom_agent = agent_kpi.iloc[-1]

fig = go.Figure()

agents_short = [a.replace(" Phillies", "") for a in agent_kpi["agent"]]
colors = [tier_color(s) for s in agent_kpi["kpi_score"]]

fig.add_trace(go.Bar(
    y=agents_short[::-1],
    x=agent_kpi["kpi_score"][::-1],
    orientation='h',
    marker_color=colors[::-1],
    text=[f"{s}%" for s in agent_kpi["kpi_score"][::-1]],
    textposition='outside',
    textfont=dict(size=12, color=C["text"], family="Arial"),
    hovertemplate="<b>%{y}</b><br>KPI Score: %{x:.1f}%<extra></extra>",
    showlegend=False,
))

# Tier zones
for x0, x1, col, label in [
    (0, 40, "rgba(239,68,68,0.08)", ""),
    (40, 60, "rgba(245,158,11,0.08)", ""),
    (60, 80, "rgba(34,197,94,0.06)", ""),
    (80, 100, "rgba(0,102,204,0.07)", ""),
]:
    fig.add_vrect(x0=x0, x1=x1, line_width=0, fillcolor=col)

fig.update_layout(
    title=dict(
        text="<b>AGS Phillies — KPI Performance Score</b><br><span style='font-size:12px;color:#64748B'>All Agents Ranked by Weighted KPI Score</span>",
        font=dict(size=18, color=C["text"], family="Arial Black"),
        x=0.5, y=0.98,
    ),
    margin=dict(l=20, r=20, t=110, b=60),
    plot_bgcolor="white",
    paper_bgcolor="white",
    height=600,
    width=1100,
    xaxis=dict(
        range=[0, 108], tickvals=[0, 20, 40, 60, 80, 100],
        ticktext=["0%", "20%", "40%", "60%", "80%", "100%"],
        showgrid=True, gridcolor="#E2E8F0",
        title="KPI Score (%)", title_font=dict(size=12, color=C["muted"]),
        tickfont=dict(size=11, color=C["muted"]),
    ),
    yaxis=dict(
        showgrid=False,
        tickfont=dict(size=12, color=C["text"], family="Arial"),
        dtick=1,
    ),
    annotations=[
        dict(x=20, y=-0.8, text="Critical <40%", showarrow=False,
             font=dict(size=9, color=C["bad"]), xanchor="center"),
        dict(x=50, y=-0.8, text="Needs Work 40-60%", showarrow=False,
             font=dict(size=9, color=C["warn"]), xanchor="center"),
        dict(x=70, y=-0.8, text="Good 60-80%", showarrow=False,
             font=dict(size=9, color=C["good"]), xanchor="center"),
        dict(x=90, y=-0.8, text="Excellent ≥80%", showarrow=False,
             font=dict(size=9, color=C["tier_excellent"]), xanchor="center"),
    ],
    uniformtext_minsize=8,
    uniformtext_mode='show',
)

# ─── Save PNG ─────────────────────────────────────────────────────────────────
output_dir = "/Users/archerterminez/Desktop/repository/ags-dashboard"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "ags_kpi_dashboard.png")

fig.write_image(output_path, format="png", width=1100, height=800, scale=2,
                 background="white")

print(f"✅ Dashboard image saved to: {output_path}")
print(f"\n📊 Data Summary:")
print(f"   Agents: {len(agent_kpi)}")
print(f"   Avg KPI Score: {avg_kpi:.1f}%")
print(f"   Top Performer: {top_agent['agent'].replace(' Phillies', '')} ({top_agent['kpi_score']}%)")
print(f"   Needs Attention: {bottom_agent['agent'].replace(' Phillies', '')} ({bottom_agent['kpi_score']}%)")
print(f"   Weeks of Data: {agent_kpi['weeks_active'].max()}")
